"""
Программа для чтения 20 канального Keysight DAQ970A
В режимах DCV и FRES

Задачи:

1) Нормально расписать документацию, а то сам забуду что и где должно быть и как работает
2) Выбор FRES выключает ch10-20 в зависимости от того на каком канале он выбран(1-10), т.к. привыставлении FRES он берет значения только с первых 10 каналов
3) Добавить больше сигналов в статус бар
4) Сделать создание Excel подстраиваемое к выбранным настройкам (выучить русский, а то тяжело предложения формулировать)
"""

import sys
import time
import os
import xlwings as xw
import pyvisa
from PyQt6.QtWidgets import QRadioButton
from PyQt6.uic import loadUi
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QPushButton, QCheckBox, QLineEdit
from PyQt6.QtCore import pyqtSignal, QThread, QSettings, QTimer

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
import os.path


class ExcelWriterThread(QThread):
    stop_signal = pyqtSignal()
    update_excel_signal = pyqtSignal(int, int, object)
    update_values_signal = pyqtSignal(int)

    def __init__(self, app_instance, parent=None):
        super().__init__(parent)
        self.app_instance = app_instance

        self.daq = self.app_instance.daq
        self.nplc_dcv = self.app_instance.nplc_dcv.text()
        self.nplc_fres = self.app_instance.nplc_fres.text()
        self.read_one_flag = self.app_instance.read_one_flag
        self.statusbar = self.app_instance.statusbar

        self.running = True
        self.excel_row = int(self.app_instance.excel_row.text())  # Начальная строка записи

    def run(self):
        current_row = self.excel_row

        while self.running:
            read_number = 1  # Для перебора каналов
            self.update_excel_signal.emit(current_row, 1, time.time() - self.app_instance.start_time)

            for i in range(1, 21):
                # ! Не знаю, будет ли грузить GUI, если да, то нужно будет что-то придумывать
                checkbox = self.app_instance.findChild(QCheckBox, f"ch{i}")
                if checkbox.isChecked():
                    dcv_radio = self.app_instance.findChild(QRadioButton, f"dcv{i}")
                    # fres_radio = self.app_instance.findChild(QRadioButton, f"fres{i}")
                    if dcv_radio.isChecked():
                        read = self.dcv_read(i, self.nplc_dcv)
                    else:
                        read = self.fres_read(i, self.nplc_fres)
                else:
                    continue

                self.update_excel_signal.emit(current_row, read_number + 1, read)

                read_number += 1

            self.update_excel_signal.emit(current_row, read_number + 1,
                                          str(time.strftime("%d.%m.%Y  %H:%M:%S", time.localtime())))

            current_row += 1  # Переход на следующую строку после записи данных
            self.update_values_signal.emit(current_row)

            if self.read_one_flag:
                self.stop()
            else:
                # Было
                # time_sleep = int(self.app_instance.time_oprosa.text())
                # time.sleep(time_sleep)

                # Стало
                time_sleep = int(self.app_instance.time_oprosa.text())
                while time_sleep > 0:
                    # print(time_sleep)
                    if time_sleep > 10:
                        self.statusbar.showMessage(f"Измерения продолжатся через {time_sleep} секунд")
                    if self.running:
                        if (time_sleep - 1) > 0:
                            time.sleep(1)  # Пауза между записями
                            time_sleep -= 1
                        else:
                            time.sleep(time_sleep)
                            time_sleep -= 1
                            break
                    else:
                        self.stop()
                        break

    def stop(self):
        self.running = False
        self.stop_signal.emit()

    def dcv_read(self, ch, nplc):
        if ch > 9:
            self.daq.write(f"CONF:VOLT:DC (@1{ch})")
        else:
            self.daq.write(f"CONF:VOLT:DC (@10{ch})")
        self.daq.write("VOLT:DC:IMP:AUTO ON")
        self.daq.write(f"VOLT:DC:NPLC {nplc}")
        temp_values = self.daq.query_ascii_values("READ?")
        return temp_values[0]

    def fres_read(self, ch, nplc):
        if ch > 9:
            self.daq.write(f"CONF:FRES (@1{ch})")
        else:
            self.daq.write(f"CONF:FRES (@10{ch})")
        self.daq.write(f"FRES:NPLC {nplc}")
        temp_values = self.daq.query_ascii_values("READ?")
        return temp_values[0]


class App(QMainWindow):
    def __init__(self):
        super().__init__()

        # self.Excel = win32.Dispatch('Excel.Application')
        # self.Excel.Visible = True  # Создаем COM объект

        self.rm = pyvisa.ResourceManager()

        loadUi('daq_reader_by_df.ui', self)  # Загружаем UI напрямую в self
        self.setWindowTitle("Excel Writer")

        self.settings = QSettings("lab425", "Reader")

        self.excel_name.setText(time.strftime("%d%m%Y ", time.localtime()) + "_Example")

        # Анимация
        self.animation_timer = QTimer()
        self.animation_timer.setInterval(500)
        self.animation_timer.timeout.connect(self.animate_text)
        self.animation_num = 2

        self.wb = None
        self.ws = None
        self.daq = None
        self.time_open = None
        self.start_time = time.time()
        self.usb_resources = None
        self.cache_values = []

        self.start_time_flag = False
        self.ip_daq_flag = True
        self.read_one_flag = False
        self.rts_flag = False  # Ready to Start
        self.start_flag = False

        self.create_excel.clicked.connect(self.create_excel_func)
        self.open_excel.clicked.connect(self.open_excel_func)
        self.start_button.clicked.connect(self.start)
        self.pause_button.clicked.connect(self.pause)
        self.connect_button.clicked.connect(self.connect)
        self.read_one.clicked.connect(self.read_one_func)
        self.save_button.clicked.connect(self.save_settings)

        self.mainthread = None

        # Переписать, а то бред какой-то тут
        # self.settings.setValue("first_run", 1)
        if self.settings.value("first_run") == 1:
            self.settings.clear()
            self.settings.sync()
            self.save_settings()
            self.settings.setValue("first_run", 2)

        self.load_settings()

    def start(self):
        if self.rts_flag:
            self.start_button.setText("Измеряется")
            self.start_flag = True
            self.disable_enable_ch()
            self.animation_timer.start()
            try:
                if self.wb is None:
                    print("Перед запуском убедитесь, что Excel создан")
                    self.statusbar.showMessage("Перед запуском убедитесь, что Excel создан")
                    self.pause()
                    return
                if self.ip_daq_flag:
                    if self.mainthread is None or not self.mainthread.isRunning():
                        self.mainthread = ExcelWriterThread(self)

                        self.mainthread.update_values_signal.connect(self.update_values)
                        self.mainthread.update_excel_signal.connect(self.update_excel)
                        self.mainthread.start()
                    else:
                        self.pause()
                else:
                    print("Poka net")
            except Exception as e:
                print(e)
                self.pause()

    def update_excel(self, row, col, value):
        try:
            for _ in self.cache_values:
                self.ws.cells(_[0], _[1]).value = _[2]
            self.cache_values.clear()

            self.ws.cells(row, col).value = value

        except:
            # [row, col]
            self.cache_values.append([row, col, value])
            print(f"Cache: {self.cache_values}")

    def update_values(self, num):
        self.excel_row.setText(f'{num}')

    def pause(self):
        self.start_flag = False
        self.disable_enable_ch()

        if self.animation_timer.isActive():
            self.animation_timer.stop()
            self.start_button.setText("Старт")
            self.animation_num = 2

        if self.mainthread and self.mainthread.isRunning():
            self.mainthread.stop()
            self.mainthread.wait()
            self.mainthread = None

        self.statusbar.showMessage("Измерение остановлено")

    def disable_enable_ch(self):
        if self.start_flag:
            for i in range(1, 21):
                current_ch = self.findChild(QCheckBox, f"ch{i}")
                current_dcv = self.findChild(QRadioButton, f"dcv{i}")
                current_fres = self.findChild(QRadioButton, f"fres{i}")
                current_ch.setEnabled(False)
                current_dcv.setEnabled(False)
                current_fres.setEnabled(False)
        else:
            for i in range(1, 21):
                current_ch = self.findChild(QCheckBox, f"ch{i}")
                current_dcv = self.findChild(QRadioButton, f"dcv{i}")
                current_fres = self.findChild(QRadioButton, f"fres{i}")
                current_ch.setEnabled(True)
                current_dcv.setEnabled(True)
                current_fres.setEnabled(True)

    def animate_text(self):
        if self.animation_num == 1:
            self.start_button.setText("Измеряется")
        elif self.animation_num == 2:
            self.start_button.setText("Измеряется.")
        elif self.animation_num == 3:
            self.start_button.setText("Измеряется..")
        else:
            self.start_button.setText("Измеряется...")
            self.animation_num = 0
        self.animation_num += 1

    def connect(self):
        try:
            if self.ip_daq_flag and self.lan.isChecked():
                self.daq = self.rm.open_resource('TCPIP0::' + str(self.ip_daq.text()) + '::inst0::INSTR')
            elif self.ip_daq_flag and self.usb.isChecked():
                self.usb_resources = [res for res in self.rm.list_resources() if res.startswith('USB')]
                print(self.usb_resources)
                # self.daq = self.rm.open_resource(self.usb_resources[0])
                # Вставить USB ниже и раскомментить, а выше закоментить если не подключается
                # self.daq = self.rm.open_resource('USB0::0x2A8D::0x5101::MY58003604::0::INSTR')  # 120
                self.daq = self.rm.open_resource('USB0::0x2A8D::0x5101::MY58038716::0::INSTR')  # 107

            self.statusbar.showMessage('Подключение успешно')
            self.rts_flag = True
            # print('Подключение успешно')
        except Exception as e:
            error = QMessageBox()
            error.setWindowTitle('Ошибка')
            error.setText(
                f'Подключение не удалось. Если подключались по usb, то проверте ip в команднике и программе(строка - 192): {e}')
            error.exec()

    def create_excel_func(self):
        try:

            file_name = str(self.excel_name.text()) + '.xlsx'

            book = Workbook()
            sheet = book.active

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            font = Font(b=True, size=12, color="000000")
            fill = PatternFill("solid", fgColor="FFFF00")

            sheet['A1'] = 'Время'
            sheet['B1'] = 'V1'
            sheet['C1'] = 'V2'
            sheet['D1'] = 'V3'
            sheet['E1'] = 'V4'
            sheet['F1'] = 'Системное время'
            sheet['G1'] = 'Комментарии Python'
            sheet['H1'] = 'Для комментариев'

            # sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['H'].width = 25
            sheet.column_dimensions['F'].width = 25
            sheet.column_dimensions['L'].width = 20
            sheet.column_dimensions['M'].width = 25
            sheet.column_dimensions['N'].width = 25

            for i in range(1, 25):
                sheet.cell(row=1, column=i).font = font
                sheet.cell(row=1, column=i).fill = fill
                sheet.cell(row=1, column=i).alignment = Alignment(horizontal='center')
                sheet.cell(row=1, column=i).border = thin_border

            if os.path.isfile("C:\\Python\\" + file_name):
                error = QMessageBox()
                error.setWindowTitle('Ошибка')
                error.setText('Такой файл уже существует, выберите другое имя')
                # error.setIcon(QMessageBox.Warning)
                # error.setStandardButtons(QMessageBox.Ok)
                error.exec()
            else:
                book.save("C:\\Python\\" + file_name)
                # os.startfile("C:\\Python\\" + file_name)
                book.close()
                self.time_open = time.time()
                while time.time() - self.time_open < 2:
                    pass
                self.wb = xw.Book(f"C:\\Python\\{file_name}")
                self.ws = self.wb.sheets[0]

        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось создать файл: {e}')

    def open_excel_func(self):
        try:
            file_name = str(self.excel_name.text()) + '.xlsx'
            self.wb = xw.Book(f"C:\\Python\\{file_name}")
            self.ws = self.wb.sheets[0]
            # xw.App(visible=True)
        except Exception as e:
            self.statusbar.showMessage(f'Ошибка.\nНе удалось открыть Excel: {e}')

    # def closeEvent(self, event):
    #     try:
    #         self.wb.close()
    #         print("Excel закрыт")
    #     except Exception as e:
    #         print(f"Ошибка при закрытии Excel: {e}")
    #     event.accept()

    def save_settings(self):
        for i in range(1, 21):
            checkbox = self.findChild(QCheckBox, f"ch{i}")
            if checkbox:
                self.settings.setValue(f"ch{i}", checkbox.isChecked())

            dcv_radio = self.findChild(QRadioButton, f"dcv{i}")
            if dcv_radio:
                self.settings.setValue(f"dcv{i}", dcv_radio.isChecked())

            fres_radio = self.findChild(QRadioButton, f"fres{i}")
            if fres_radio:
                self.settings.setValue(f"fres{i}", fres_radio.isChecked())

        self.settings.setValue("ip_daq", self.ip_daq.text())
        self.settings.setValue("time_oprosa", self.time_oprosa.text())
        self.settings.setValue("nplc_dcv", self.nplc_dcv.text())
        self.settings.setValue("nplc_fres", self.nplc_fres.text())

        self.statusbar.showMessage('Настройки сохранены')

    def load_settings(self):
        for i in range(1, 21):
            checkbox = self.findChild(QCheckBox, f"ch{i}")
            if checkbox:
                checkbox.setChecked(self.settings.value(f"ch{i}", False, type=bool))

            dcv_radio = self.findChild(QRadioButton, f"dcv{i}")
            fres_radio = self.findChild(QRadioButton, f"fres{i}")

            if dcv_radio:
                dcv_checked = self.settings.value(f"dcv{i}", False, type=bool)
                dcv_radio.blockSignals(True)
                dcv_radio.setChecked(dcv_checked)
                dcv_radio.blockSignals(False)

            if fres_radio:
                fres_checked = self.settings.value(f"fres{i}", False, type=bool)
                fres_radio.blockSignals(True)
                fres_radio.setChecked(fres_checked)
                fres_radio.blockSignals(False)

        self.ip_daq.setText(self.settings.value("ip_daq", ""))
        self.time_oprosa.setText(self.settings.value("time_oprosa", ""))
        self.nplc_dcv.setText(self.settings.value("nplc_dcv", ""))
        self.nplc_fres.setText(self.settings.value("nplc_fres", ""))

    def read_one_func(self):
        # print([child.objectName() for child in self.findChildren(QRadioButton)])
        self.read_one_flag = True
        self.start()
        self.read_one_flag = False


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = App()
    window.show()
    sys.exit(app.exec())
